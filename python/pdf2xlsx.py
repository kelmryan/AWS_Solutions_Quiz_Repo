import PyPDF2 as p2
import xlsxwriter
import sys
import openpyxl

def extract_relevant_info(rawInfo):
    start_index = None
    end_index = None
    relevant_info = []

    for i, line in enumerate(rawInfo):
        if line.startswith("Topic 1 Question"):
            start_index = i
        if line.startswith("Correct Answer:") and start_index is not None:
            end_index = i + 1
            relevant_info.extend(rawInfo[start_index:end_index])
            start_index = None  # Reset start_index to look for the next "Topic 1 Question"
    
    return relevant_info

def pdf_to_excel(pdfFileName):
    pdfFile = open(pdfFileName, 'rb')
    pdfread = p2.PdfFileReader(pdfFile)
    number_of_pages = pdfread.getNumPages()
    workbook = xlsxwriter.Workbook('pdftoexcel.xlsx')

    for page_number in range(number_of_pages):
        pageinfo = pdfread.getPage(page_number)
        rawInfo = pageinfo.extractText().split('\n')
        relevant_info = extract_relevant_info(rawInfo)

        row = 0
        column = 0
        worksheet = workbook.add_worksheet(f'Sheet{page_number}')

        for line in relevant_info:
            worksheet.write(row, column, line)
            row += 1

    workbook.close()

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python pdf2xlsx.py <pdf_file>")
    else:
        pdfFileName = sys.argv[1]
        pdf_to_excel(pdfFileName)
        def remove_blank_sheets(workbook_path):

            workbook = openpyxl.load_workbook(workbook_path)
            sheets_to_remove = []

            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet.cell(1, 1).value is None:
                    sheets_to_remove.append(sheet)

            for sheet in sheets_to_remove:
                workbook.remove(workbook[sheet])

            workbook.save(workbook_path)

        if __name__ == "__main__":
            if len(sys.argv) != 2:
                print("Usage: python pdf2xlsx.py <pdf_file>")
            else:
                pdfFileName = sys.argv[1]
                pdf_to_excel(pdfFileName)
                remove_blank_sheets('pdftoexcel.xlsx')